import argparse
from setup import *
import pandas as pd
from openpyxl import load_workbook

def parse_args():
    """
    Notes:
        Parse command line arguments.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument("--file_date", type=str, required=True)
    parser.add_argument("--update", action='store_true')
    return parser.parse_args()


def get_manual_store_names(source_manual_file):
    """
    Notes:
        Get store names from store_info.xlsx.
    
    Args:
        source_manual_file (str): source manual file.
    
    Returns:
        list: store names.
    """
    routes_df = pd.read_excel(source_manual_file, sheet_name=0, skiprows=3)
    store_names = routes_df['店名'].tolist()
    store_names = [name.strip() for name in store_names if name != '林口ＤＣ']
    return store_names


def get_store_info_name(store_info_file):
    """
    Notes:
        Get store names from store_info.xlsx.
    
    Args:
        store_info_file (str): store info file.
    
    Returns:
        list: store names.
    """
    store_info_df = pd.read_excel(store_info_file, sheet_name=0)
    store_names = store_info_df['店鋪名稱'].tolist()
    store_names = [name.strip() for name in store_names]
    return store_names


def get_store_info(store_info_file, stores):
    """
    Notes:
        Get store info from store_info.xlsx.
    
    Args:
        store_info_file (str): store info file.
    
    Returns:
        dict: store info.
    """
    store_info = []
    columns = ['店鋪編號', 'DC別', '店鋪名稱', '店鋪地址', '經度', '緯度', '營業狀態', '夜間熄火', '不可大車', '不可爆量', '不可進店時段1', '不可進店時段2', '不可進店時段3']
    store_info_df = pd.read_excel(store_info_file, sheet_name=0)

    for store_name in stores:
        filtered_df = store_info_df[store_info_df['店鋪名稱'] == store_name]
        if filtered_df.empty:
            continue
        first_row = filtered_df.iloc[0][columns]
        store_info.append(first_row.to_dict())

    return store_info


def add_store_info(store_info_file, store_info):
    """
    Notes:
        Add store info to store_info.xlsx.
    
    Args:
        store_info_file (str): store info file.
        store_info (list): store info.
    
    Returns:
        None
    """
    if not store_info:
        print("No new store info to add.")
        return

    columns_needed = ['店鋪編號', 'DC別', '店鋪名稱', '店鋪地址', '經度', '緯度', '營業狀態', '夜間熄火', '不可大車', '不可爆量', '不可進店時段1', '不可進店時段2', '不可進店時段3']

    new_df = pd.DataFrame([{k: store.get(k, None) for k in columns_needed} for store in store_info])
    file_path = Path(store_info_file)

    if not file_path.exists():
        new_df.to_excel(store_info_file, index=False)
        print(f"Created new file and added {len(new_df)} stores.")
        return

    with pd.ExcelWriter(store_info_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        sheet_name = writer.book.sheetnames[0]
        startrow = writer.book[sheet_name].max_row
        new_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)

    print(f"Added {len(new_df)} stores to {store_info_file}.")


def main(file_date, update=False):
    """
    Notes:
        Transfer manual data to origin data.
    """
    source_manual_file = f'../data/manual_data/{file_date}manual.xlsx'
    new_store_info_file = '../data/new_store_info.xlsx'
    store_info_file = '../data/store_info.xlsx'

    manual_store_names = get_manual_store_names(source_manual_file)
    store_info_names = get_store_info_name(store_info_file)

    new_store = []
    for name in manual_store_names:
        if name not in store_info_names:
            new_store.append(name)
    store_info = get_store_info(new_store_info_file, new_store)
    add_store_info(store_info_file, store_info)

    if update:
        distance_matrix, time_matrix = s_data.get_cost_matrices(dist_file, time_file)


if __name__ == "__main__":
    args = parse_args()
    main(args.file_date, args.update)

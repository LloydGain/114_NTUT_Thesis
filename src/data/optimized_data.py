import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

class OPDataManager:
    """
    Notes:
        Optimized Data Manager.
    """
    def __init__(self, source_file):
        self.source_file = source_file
        self.routes = self._load_routes()


    def _load_routes(self):
        """
        Notes:
            Load routes from JSON file.

        Args:
            None.

        Returns:
            dict: Routes data.
        """
        with open(self.source_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        return data


    def _is_within_time_window(self, arrival_time, earliest_time, latest_time):
        """
        Notes:
            Check if a given arrival time within store time window.

        Args:
            arrival_time (datetime): Arrival time.
            earliest_time (datetime): Earliest time (start of time window).
            latest_time (datetime): Latest time (end of time window).

        Returns:
            bool: True if within time window, False otherwise.
        """
        return earliest_time <= arrival_time <= latest_time


    def _summarize_routes(self, routes):
        """
        Notes:
            Generate a summary of the routes.

        Args:
            routes (dict): Routes information.

        Returns:
            dict: Summary statistics.
        """
        total_vehicles = len(routes)
        main_vehicles = sum(1 for route in routes.values() if not route['dc']['route_id'].isdigit())
        support_vehicles = sum(1 for route in routes.values() if route['dc']['route_id'].isdigit())

        total_stores = sum(len(route['stores']) for route in routes.values())
        main_stores = sum(len(route['stores']) for route in routes.values() if not route['dc']['route_id'].isdigit())
        support_stores = sum(len(route['stores']) for route in routes.values() if route['dc']['route_id'].isdigit())

        total_distance = sum(route['dc']['distance'] for route in routes.values())
        total_duration = sum(route['dc']['duration'] for route in routes.values()) / (60 * 60)
        total_main_distance = sum(route['dc']['distance'] for route in routes.values() if not route['dc']['route_id'].isdigit())
        total_main_duration = sum(route['dc']['duration'] for route in routes.values() if not route['dc']['route_id'].isdigit()) / (60 * 60)
        total_support_distance = sum(route['dc']['distance'] for route in routes.values() if route['dc']['route_id'].isdigit())
        total_support_duration = sum(route['dc']['duration'] for route in routes.values() if route['dc']['route_id'].isdigit()) / (60 * 60)

        average_distance = total_distance / total_vehicles
        average_duration = total_duration / total_vehicles
        average_main_distance = total_main_distance / main_vehicles
        average_main_duration = total_main_duration / main_vehicles
        average_support_distance = total_support_distance / support_vehicles if not support_vehicles == 0 else 0
        average_support_duration = total_support_duration / support_vehicles if not support_vehicles == 0 else 0

        average_load_rate = sum(route['dc']['load_rate'] for route in routes.values()) / total_vehicles
        average_main_load_rate = sum(route['dc']['load_rate'] for route in routes.values() if not route['dc']['route_id'].isdigit()) / main_vehicles
        average_support_load_rate = sum(route['dc']['load_rate'] for route in routes.values() if route['dc']['route_id'].isdigit()) / support_vehicles if not support_vehicles == 0 else 0

        on_time = 0
        main_on_time = 0
        support_on_time = 0
        for route in routes.values():
            stores = route['stores']
            if route['dc']['route_id'].isdigit():
                for store in stores:
                    if self._is_within_time_window(store['pred_time'], store['earliest_time'], store['latest_time']):
                        support_on_time += 1
                        on_time += 1
            else:
                for store in stores:
                    if self._is_within_time_window(store['pred_time'], store['earliest_time'], store['latest_time']):
                        main_on_time += 1
                        on_time += 1
        on_time_rate = on_time / total_stores
        main_on_time_rate = main_on_time / main_stores
        support_on_time_rate = support_on_time / support_stores if not support_stores == 0 else 0

        summary = {
            '車輛數量': {
                'main': f'{main_vehicles}',
                'support': f'{support_vehicles}',
                'total': f'{total_vehicles}',
            },
            '店數': {
                'main': f'{main_stores}',
                'support': f'{support_stores}',
                'total': f'{total_stores}',
            },
            '總路線距離': {
                'main': f'{total_main_distance:.2f}',
                'support': f'{total_support_distance:.2f}',
                'total': f'{total_distance:.2f}',
            },
            '總路線時長': {
                'main': f'{total_main_duration:.2f}',
                'support': f'{total_support_duration:.2f}',
                'total': f'{total_duration:.2f}',
            },
            '路線平均距離': {
                'main': f'{average_main_distance:.2f}',
                'support': f'{average_support_distance:.2f}',
                'total': f'{average_distance:.2f}',
            },
            '路線平均時長': {
                'main': f'{average_main_duration:.2f}',
                'support': f'{average_support_duration:.2f}',
                'total': f'{average_duration:.2f}',
            },
            '路線平均載裝載率': {
                'main': f'{average_main_load_rate:.2f}',
                'support': f'{average_support_load_rate:.2f}',
                'total': f'{average_load_rate:.2f}',
            },
            '店鋪準點率': {
                'main': f'{main_on_time_rate:.2f}',
                'support': f'{support_on_time_rate:.2f}',
                'total': f'{on_time_rate:.2f}',
            }
        }

        return summary


    def _prepare_data(self):
        """
        Notes:
            Prepare data for export.

        Args:
            None.

        Returns:
            list: List of data.
        """
        data = []
        for route in self.routes.values():
            dc = route['dc']
            stores = route['stores']

            data.append({
                "車次": dc["route_code"],
                "ID": "",
                "店名": "林口DC",
                "表定時間": "",
                "預定時間": "",
                "貨量": dc["total_volume"],
                "裝載率": dc["load_rate"],
                "總距離(km)": dc["distance"],
                "總時長(hr)": dc["duration"] / 3600,
            })

            for s in stores:
                data.append({
                    "車次": s["route_code"],
                    "ID": s["store_id"],
                    "店名": s["store_name"],
                    "表定時間": datetime.strptime(s["sched_time"], "%Y-%m-%dT%H:%M:%S").strftime("%Y-%m-%d %H:%M:%S"),
                    "預定時間": datetime.strptime(s["pred_time"], "%Y-%m-%dT%H:%M:%S").strftime("%Y-%m-%d %H:%M:%S"),
                    "貨量": s["volume"],
                    "裝載率": "",
                    "總距離(km)": "",
                    "總時長(hr)": "",
                })

            data.append({
                "車次": f'{dc["route_code"]}DC',
                "ID": "",
                "店名": "林口DC",
                "表定時間": "",
                "預定時間": "",
                "貨量": 0,
                "裝載率": "",
                "總距離(km)": "",
                "總時長(hr)": "",
            })

        return data


    def _apply_style(self, cell, bold=False):
        font_style = Font(name='Microsoft JhengHei', size=12, bold=bold)
        alignment = Alignment(horizontal='center', vertical='center')
        thin_side = Side(border_style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        cell.font = font_style
        cell.alignment = alignment
        cell.border = border


    def export_origin_excel_file(self, dest_file, sheet_name):
        """
        Notes:
            Export origin excel.

        Args:
            dest_file (str): original file.

        Returns:
            None.
        """
        data = self._prepare_data()

        columns = [
            "車次", "ID", "店名", "表定時間", "預定時間",
            "貨量", "裝載率", "總距離(km)", "總時長(hr)"
        ]

        col_widths = [20, 20, 20, 28, 28, 15, 20, 18, 18, 10, 10, 24, 15, 15, 15]
        df = pd.DataFrame(data, columns=columns)

        mode = "a" if os.path.exists(dest_file) else "w"

        with pd.ExcelWriter(dest_file, engine="openpyxl", mode=mode) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(dest_file)
        ws = wb[sheet_name]

        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                if row_idx == 1:
                    self._apply_style(cell, bold=True)
                else:
                    self._apply_style(cell)

        summary = self._summarize_routes(self.routes)

        cell = ws.cell(row=2, column=len(columns) + 3)
        cell.value = ""
        self._apply_style(cell)

        cell = ws.cell(row=2, column=len(columns) + 4)
        cell.value = "主線"
        self._apply_style(cell)

        cell = ws.cell(row=2, column=len(columns) + 5)
        cell.value = "支援線"
        self._apply_style(cell)

        cell = ws.cell(row=2, column=len(columns) + 6)
        cell.value = "總計"
        self._apply_style(cell)

        for i, (key, value) in enumerate(summary.items(), start=1):
            cell = ws.cell(row=i + 2, column=len(columns) + 3)
            cell.value = key
            self._apply_style(cell)
            for j, (_, sub_value) in enumerate(value.items(), start=1):
                cell = ws.cell(row=i + 2, column=len(columns) + j + 3)
                cell.value = sub_value
                self._apply_style(cell)

        wb.save(dest_file)
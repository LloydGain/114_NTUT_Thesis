import os
import copy
import shutil
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from models.route_manager import RouteManager
from data.base_data import BaseDataManager

class MDataManager(BaseDataManager):
    """
    Notes:
        Load route-related data from source files (manual).
    """
    def __init__(self, excel_files, distance_matrix=None, time_matrix=None):
        super().__init__(excel_files, distance_matrix, time_matrix)
        self.routes_info = self._load_manual_routes()
        if distance_matrix is not None and time_matrix is not None:
            self._update_routes_info()


    def _load_manual_routes(self):
        """
        Notes:
            Load routes data from excel & transfer to dict.

        Args:
            file (str): Path to the Excel File.

        Returns:
            dict: routes information, with route code as key.
        """
        routes_info = {}

        update_main_route = True
        main_route_code = None
        for _, row in self.routes_df.iterrows():
            route_code = str(row['車次'])

            if not route_code:
                continue

            store_name = row['店名']
            store_id = self._get_store_id_by_name(store_name)

            if store_id is None and ('ＤＣ' not in store_name and '專車' not in store_name):
                raise ValueError(f"Store ID not found for store name: {store_name}")

            lng, lat = self._get_coordinates(store_id)
            dwell_time = self._get_dwell_time(store_id)
            sched_time = pd.to_datetime(row['表定時間']).isoformat()
            earliest_time = (pd.to_datetime(row['表定時間']) - pd.Timedelta(hours=1)).isoformat()
            latest_time = (pd.to_datetime(row['表定時間']) + pd.Timedelta(minutes=30)).isoformat()
            pred_time = pd.to_datetime(row['預定時間']).isoformat()
            volume = row['貨量']
            load_rate = row['裝載率']

            if main_route_code is None:
                main_route_code = route_code[:3] if route_code.isdigit() else route_code[:2]

            if update_main_route:
                main_route_code = route_code[:3] if route_code.isdigit() else route_code[:2]
                update_main_route = False

            if self._DC_CENTER in route_code:
                update_main_route = True

            max_capacity = self._get_max_capacity_by_route_code(main_route_code)

            if main_route_code not in routes_info:
                routes_info[main_route_code] = {"dc" : None, "stores": []}

            if self._DC_CENTER not in route_code and len(route_code) < 4:
                routes_info[main_route_code]["dc"] = {
                    "route_id": main_route_code,
                    "route_code": route_code,
                    "store_id": "DC",
                    "store_name": store_name,
                    "total_volume": volume,
                    "load_rate": load_rate,
                    "max_capacity": max_capacity,
                    "distance": 0,
                    "duration": 0
                }
            elif self._DC_CENTER not in route_code:
                routes_info[main_route_code]["stores"].append({
                    "route_id": main_route_code,
                    "route_code": route_code,
                    "store_id": store_id,
                    "store_name": store_name,
                    "longitude": lng,
                    "latitude": lat,
                    "sched_time": sched_time,
                    "earliest_time": earliest_time,
                    "latest_time": latest_time,
                    "pred_time": pred_time,
                    "dwell_time": dwell_time,
                    "volume": volume
                })

        return routes_info


    def _is_within_time_window(self, arrival_time, earliest_time, latest_time):
        """
        Notes:
            Check if a given arrival time within store time window.

        Args:
            arrival_time (datetime): Arrival time.
            earliest_time (datetime): Earliest time (start of time window).
            latest_time (datetime): Latest time (end of time window).

        Returns:
            bool: True if within time window, False otherwise.
        """
        return earliest_time <= arrival_time <= latest_time


    def get_invalid_routes(self):
        """
        Notes:
            Get invalid routes.

        Args:
            None.

        Returns:
            data (list): data row with routes info.
        """
        data = []
        for route_id in self.routes_info:
            dc = self.routes_info[route_id]['dc']
            stores = self.routes_info[route_id]['stores']
            for s in stores:
                if s['route_code'][:2] == dc['route_code']:
                    continue
                earliest_time = pd.to_datetime(s['earliest_time'])
                latest_time = pd.to_datetime(s['latest_time'])
                pred_time = pd.to_datetime(s['pred_time'])
                if not self._is_within_time_window(pred_time, earliest_time, latest_time):
                    data.append(self.routes_info[route_id])
                    break

        return data


    def create_data_folder(self, dst_dir, source_manual_file, dest_manual_file):
        """
        Notes:
            Create the data foler.

        Args:
            dst_dir (str): folder.
            source_manual_file (str): Source Manual file.
            dest_manual_file (str): Destination Manual file.

        Returns:
            None.
        """
        os.makedirs(dst_dir, exist_ok=True)
        shutil.copy(source_manual_file, dest_manual_file)


    def export_invalid_routes(self, dest_file, sheet_name):
        """
        Notes:
            Export invalid routes to Excel.
            - 標出不在時窗的行 (黃色)
            - 計算最早、最晚抵達時間
            - 計算時間差，顯示提前/延遲分鐘
        """
        data = []

        for route in self.get_invalid_routes():
            dc = route['dc']
            stores = route['stores']

            data.append({
                "不可大車": "",
                "車次": dc["route_code"],
                "ID": "",
                "店名": "林口DC",
                "表定時間": "",
                "預定時間": "",
                "貨量": dc["total_volume"],
                "裝載率": dc["load_rate"],
                "最早抵達時間": "",
                "最晚抵達時間": "",
                "時間差": ""
            })

            for s in stores:
                data.append({
                    "不可大車": False,
                    "車次": s["route_code"],
                    "ID": s["store_id"],
                    "店名": s["store_name"],
                    "表定時間": s["sched_time"],
                    "預定時間": s["pred_time"],
                    "貨量": s["volume"],
                    "裝載率": "",
                    "最早抵達時間": s["earliest_time"],
                    "最晚抵達時間": s["latest_time"],
                    "時間差": ""
                })

            data.append({
                "不可大車": False,
                "車次": f'{dc["route_code"]}DC',
                "ID": "",
                "店名": "林口DC",
                "表定時間": "",
                "預定時間": "",
                "貨量": 0,
                "裝載率": "",
                "最早抵達時間": "",
                "最晚抵達時間": "",
                "時間差": ""
            })

        columns = [
            "不可大車",
            "車次",
            "ID",
            "店名",
            "表定時間",
            "預定時間",
            "貨量",
            "裝載率",
            "最早抵達時間",
            "最晚抵達時間",
            "時間差"
        ]

        col_widths = [10, 8, 12, 15, 20, 20, 10, 10, 20, 20, 12]
        df = pd.DataFrame(data, columns=columns)

        mode = "a" if os.path.exists(dest_file) else "w"

        with pd.ExcelWriter(dest_file, engine="openpyxl", mode=mode) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(dest_file)
        ws = wb[sheet_name]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            sched_val = row[4].value
            pred_val = row[5].value

            if not sched_val or not pred_val:
                continue

            sched_time = datetime.fromisoformat(sched_val) if isinstance(sched_val, str) else sched_val
            pred_time = datetime.fromisoformat(pred_val) if isinstance(pred_val, str) else pred_val
            earliest_window = sched_time - timedelta(hours=1)
            latest_window = sched_time + timedelta(minutes=30)
            if not self._is_within_time_window(pred_time, earliest_window, latest_window):
                for cell in row:
                    cell.fill = yellow_fill

                time_diff = pred_time - sched_time
                minutes_diff = int(time_diff.total_seconds() / 60)

                if minutes_diff > 30:
                    minutes_diff -= 30
                elif minutes_diff < -60:
                    minutes_diff += 60

                if minutes_diff < 0:
                    row[10].value = f'早{abs(minutes_diff)}分鐘'
                else:
                    row[10].value = f'晚{minutes_diff}分鐘'

        wb.save(dest_file)


    def _get_origin_routes(self):
        """
        Note:
            Transfer the manual route to the original route.

        Args:
            None.

        Returns:
            data (list): data row with routes info.
        """
        data = []
        routes = copy.deepcopy(self.routes_info)
        route_manager = RouteManager(routes, self.distance_matrix, self.time_matrix)
        route_manager.move_store_to_original_route()
        routes = route_manager.routes_info

        for route_id in routes:
            dc = routes[route_id]['dc']
            stores = routes[route_id]['stores']
            data.append({
                "不可大車": "",
                "車次": dc["route_code"],
                "ID": "",
                "店名": '林口DC',
                "表定時間": "",
                "預定時間": "",
                "貨量": dc["total_volume"],
                "裝載率": dc["load_rate"]
            })
            for s in stores:
                data.append({
                    "不可大車": False,
                    "車次": s["route_code"],
                    "ID": s["store_id"],
                    "店名": s["store_name"],
                    "表定時間": s["sched_time"],
                    "預定時間": s["sched_time"],
                    "貨量": s["volume"],
                    "裝載率": ""
                })
            data.append({
                "不可大車": False,
                "車次": f'{dc["route_code"]}DC',
                "ID": "",
                "店名": '林口DC',
                "表定時間": "",
                "預定時間": "",
                "貨量": 0,
                "裝載率": ""
            })

        return data


    def export_origin_excel_file(self, dest_file):
        """
        Notes:
            Export origin excel.

        Args:
            dest_file (str): original file.

        Returns:
            None.
        """
        data = self._get_origin_routes()
        columns = [
            "不可大車",
            "車次",
            "ID",
            "店名",
            "表定時間",
            "預定時間",
            "貨量",
            "裝載率"
        ]
        col_widths = [10, 8, 12, 15, 20, 20, 10, 10, 20, 20, 12]

        df = pd.DataFrame(data, columns=columns)
        with pd.ExcelWriter(dest_file, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=3)

        wb = load_workbook(dest_file)
        ws = wb['Sheet1']
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(dest_file)
